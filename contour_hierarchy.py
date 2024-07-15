import os 
import openpyxl
import SimpleITK as sitk 
import numpy as np
import cv2 
order_xl_path = r"G:\! project\2024- ProstateSegmentation\20240125\data_matching_MRN-order.xlsx"
im_basepath = r"G:\! project\2024- ProstateSegmentation\20240125\savepath_mr"
la_basepath = r"G:\! project\2024- ProstateSegmentation\20240125\GT001_20240124_draft_rev"
rv_basepath = r"G:\! project\2024- ProstateSegmentation\20240125\GT002_20240200_AAPM_MR"
savepath =r"G:\! project\2024- ProstateSegmentation\20240125\RTST_contourpoint_npy"

im_dirs_list = sorted(os.listdir(im_basepath))
la_dirs_list = sorted(os.listdir(la_basepath))
rv_dirs_list = sorted(os.listdir(rv_basepath))
wb = openpyxl.load_workbook(order_xl_path)
ws = wb["Sheet"]

for idx in range(len(im_dirs_list)):
    MRN = str(ws.cell(idx+1, 1).value)
    PATIENT = str(ws.cell(idx+1, 2).value)
    im_fol_name = im_dirs_list[idx]
    la_name = im_fol_name + "_label.nii.gz"
    label_path = os.path.join(la_basepath, la_name)
    print(label_path)
    if im_fol_name in rv_dirs_list:
        la_name = os.listdir(os.path.join(rv_basepath, im_fol_name))[0]
        la_name = os.listdir(os.path.join(rv_basepath, im_fol_name))[0]
        label_path = os.path.join(rv_basepath, im_fol_name, la_name)
    print(label_path)
    image_path = os.path.join(im_basepath, im_dirs_list[idx], im_dirs_list[idx] + "_image.nii.gz")
    
    image_itk = sitk.ReadImage(image_path)
    label_itk = sitk.ReadImage(label_path)
    image_arr = sitk.GetArrayFromImage(image_itk)
    label_arr = sitk.GetArrayFromImage(label_itk)
    image_arr = np.uint8(255 * (image_arr.astype("float64") - np.min(image_arr))/(2000 - np.min(image_arr)))

    origin = np.array(image_itk.GetOrigin())
    spacing = np.array(image_itk.GetSpacing())
    contour_dict = {1: [], 2: [], 3: [], 5: []}

    for i in range(len(label_arr)):
        
        for oar in range(1, 6): 
            if oar == 4:
                continue
            
            if np.any(label_arr[i] == oar):
                binary = np.where(label_arr[i] == oar, 1, 0).astype("uint8")
                contours_co, hierarchy = cv2.findContours(binary, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                co_total = None
                for contours_sep in contours_co:
                    co = np.transpose(np.squeeze(contours_sep)) # x, y
                    # co = np.flip(co, axis = 0) # y, x
                    if co.ndim == 1:
                        continue
                    z_co = np.zeros((1, len(co[0]))).astype("int32")
                    z_co += i
                    co = np.concatenate((co, z_co), axis = 0)
                    co = np.transpose(co)
                    if co_total is not None:
                        co_total = np.concatenate((co_total, co), axis = 0)
                    else:
                        co_total = np.copy(co)
                if co_total is None:
                    continue
                
                # co_total = co_total * spacing  + origin 
                contour_dict[oar].append(co_total.astype("int32"))
    # for key,val in contour_dict.items():
    #     val = np.array(val).astype("int32")
            
    np.save(os.path.join(savepath, MRN+ "_" + PATIENT +".npy"), contour_dict)
                

            
